###
# Simply flask app hosting company data such as dep.html, name.html and so on
##

from flask import Flask, redirect, url_for, render_template, request, send_file
import openpyxl
from openpyxl import Workbook
import csv
import xlrd
import subprocess
import pandas as pd

app = Flask(__name__)

def zapisz_pliki(file_pracownicy,file_planowanie,file_pracownicydict,file_projekty,file_projektydict,file_specjalnosci):
	wb = openpyxl.load_workbook(f'excels/{file_pracownicy}')
	wb2 = openpyxl.load_workbook(f'excels/{file_planowanie}')
	wb3 = openpyxl.load_workbook(f'excels/{file_pracownicydict}')
	wb4 = openpyxl.load_workbook(f'excels/{file_projekty}')
	wb5 = openpyxl.load_workbook(f'excels/{file_projektydict}')
	wb6 = openpyxl.load_workbook(f'excels/{file_specjalnosci}')
	sh = wb.active
	sh2 = wb2.active
	sh3 = wb3.active
	sh4 = wb4.active
	sh5 = wb5.active
	sh6 = wb6.active

	with open('csvs/pracownicy.csv', 'w', newline="") as f:  
	    c = csv.writer(f)
	    for r in sh.rows:
	        c.writerow([cell.value for cell in r])

	with open('csvs/planowanie.csv', 'w', newline="") as f2:  
	    c2 = csv.writer(f2)
	    for r in sh2.rows:
	        c2.writerow([cell.value for cell in r])

	with open('csvs/pracownicy_dict.csv', 'w', newline="") as f:  
	    c3 = csv.writer(f)
	    for r in sh3.rows:
	        c3.writerow([cell.value for cell in r])

	with open('csvs/projekty.csv', 'w', newline="") as f:  
	    c4 = csv.writer(f)
	    for r in sh4.rows:
	        c4.writerow([cell.value for cell in r])

	with open('csvs/projektydict.csv', 'w', newline="") as f:  
	    c5 = csv.writer(f)
	    for r in sh5.rows:
	        c5.writerow([cell.value for cell in r])

	with open('csvs/specjalnosci.csv', 'w', newline="") as f:  
	    c6 = csv.writer(f)
	    for r in sh6.rows:
	        c6.writerow([cell.value for cell in r])

	# Tworzy plik planowanie
	subprocess.call("Rscript run.R", shell=True)

	wb_planowanie = Workbook()
	ws_planowanie = wb_planowanie.active

	with open('csvs/planowanie.csv', 'r') as f:
		for row in csv.reader(f):
			ws_planowanie.append(row)
	wb_planowanie.save('excels/planowanie.xlsx')

@app.route("/")
def home():
	return render_template("index.html")

@app.route("/pliki", methods=["POST", "GET"])
def pliki():
	if request.method == "POST":
		file_pracownicy = request.form["file_pracownicy"]
		file_planowanie = request.form["file_planowanie"]
		file_pracownicydict = request.form["pracownicydict"]
		file_projekty = request.form["projekty"]
		file_projektydict = request.form["projektydict"]
		file_specjalnosci = request.form["specjalnosci"]
		zapisz_pliki(file_pracownicy,file_planowanie,file_pracownicydict,file_projekty,file_projektydict,file_specjalnosci)
		return render_template("zaladowano_pliki.html")

	return render_template("pliki.html")

@app.route("/pracownicy")
def pracownicy():
	return render_template("pracownicy.html")

@app.route("/planowanie")
def planowanie():
	return render_template("planowanie.html")

@app.route("/podsumowanie_osobowe")
def pod_os():
	return render_template("podsumowanie.html")

@app.route("/pracownicy_podglad")
def prac_view():
	pracownicy_df = pd.read_csv('csvs/pracownicy.csv')
	pracownicy_df = pracownicy_df.fillna("")
	return render_template( 'simple.html',  tables=[pracownicy_df.to_html(classes='data')], titles=pracownicy_df.columns.values ) 

@app.route("/planowaniee_podglad")
def plan_view():
	planowanie_df = pd.read_csv('csvs/planowanie.csv')
	planowanie_df = planowanie_df.fillna("")
	return render_template( 'simple.html',  tables=[planowanie_df.to_html(classes='data')], titles=planowanie_df.columns.values ) 

@app.route("/planowwaniee_download")
def plan_down():
	return send_file("excels/planowanie.xlsx",as_attachment=True)

@app.route("/pracownicyy_download")
def prac_down():
	return send_file("csvs/pracownicy.xlsx",as_attachment=True)

@app.route("/planowanie_forms")
def plan_form():
	return render_template("planowanie.html")

if __name__ == "__main__":
	app.run()


