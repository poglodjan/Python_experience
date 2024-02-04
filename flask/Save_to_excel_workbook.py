import pandas as pd
import subprocess
from os import listdir
from os.path import isfile, join
import xlwings as xw
import os 
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


wb = Workbook()
ws = wb.create_sheet('eg')

data = [('id', 'name', 'country'),
        (1, 'Trudeau','Canada'),
        (2,'Zelenskyy', 'Ukraine'),
        (3,'Putin', 'Russia'),
        (4,),
        (5,'Biden', 'US'),
        (6, 'Xi', 'China'),
        (7,'Johnson', 'UK'),
        (8, 'Castex', 'France'),
        (9,),
        (10, 'Steinmeier', 'Germany'),
        (11, 'Rutte', 'Netherlands'),
        (12,'Loong', 'Singapore')]

for row in data:
    ws.append(row)

ss = openpyxl.load_workbook("result.xlsx")
ss_sheet = ss['edor']

fill_cell = PatternFill(patternType='solid', fgColor='c9e1f8') 
for cell in ss["2:2"]:
    cell.fill = fill_cell

file = 'result.xlsx'
wb = openpyxl.load_workbook(filename=file)
ws = wb['edor']

nrow = ws.max_row
for i in range(1,nrow):
    if i%2!=0:
        fill_cell = PatternFill(patternType='solid', fgColor='c9e1f8') 
    else:
        fill_cell = PatternFill(patternType='solid', fgColor='1E90FF') 
    for cell in ws[f"{i}:{i}"]:
        cell.fill = fill_cell


nazwy=['edor', 'portal rp', 'pua', 'Worklogall', 'nienazwane projekty', 'kap', 'ecmentarz']
ss=openpyxl.load_workbook("result.xlsx")
sheets = ss.sheetnames
s=len(sheets)

pom=sheets[0]
print(pom)
for i in range(s):
    ss_sheet = ss[f'{sheets[i]}']
    ss_sheet.title = f'{nazwy[i]}'
    ss.save("result.xlsx")
   
for i in range(s):
    ss_sheet = ss[f'{sheets[i]}']
    ss_sheet = ss[f'{nazwy[i]}']
    nrow = ss_sheet.max_row
    for j in range(1,nrow+1):
        if j%2==0:
            fill_cell = PatternFill(patternType='solid', fgColor='c9e1f8') 
        else:
            fill_cell = PatternFill(patternType='solid', fgColor='1E90FF') 
        for cell in ss_sheet[f"{j}:{j}"]:
            cell.fill = fill_cell


ss=openpyxl.load_workbook(f"{k}")
sheets = ss.sheetnames
for i in range(2):
    ss_sheet = ss[f'{sheets[i]}']
    if ss_sheet.title != 'Worklog':
        ss.remove_sheet(sheets[i])
    ss.save(f"{k}")

sheets = ss.sheetnames
s=len(sheets)
for i in range(s):
    ss_sheet = ss[f'{sheets[i]}']
    if sheets[i] != "Worklog":
        std = ss.get_sheet_by_name(f'{sheets[i]}')
        ss.remove_sheet(std)
    ss.save(f"{k}")


