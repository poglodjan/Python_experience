import pandas as pd
import subprocess
from os import listdir
from os.path import isfile, join
import xlwings as xw
import os 
import openpyxl
from openpyxl.styles import PatternFill

#   Szuka pliku pobranego z jiry 
def find():
    onlyfiles = [f for f in listdir("./") if isfile(join("./",f))]
    n=len(onlyfiles)
    for i in range(n):
        file = onlyfiles[i]
        if file[-1] == "x":
            ex_file = file
            return ex_file
        else: continue

def is_csv():
    table = []
    onlyfiles = [f for f in listdir("./") if isfile(join("./",f))]
    n=len(onlyfiles)
    for i in range(n):
        file = onlyfiles[i]
        if file[-1] == "v":
            table.append(file)
        else: continue
    return table

def is_xlsx():
    table = []
    onlyfiles = [f for f in listdir("./") if isfile(join("./",f))]
    n=len(onlyfiles)
    for i in range(n):
        file = onlyfiles[i]
        if file[-1] == "x":
            table.append(file)
        else: continue
    return table

# zapisuje z csv do xlsx
def to_xlsx(onlyfiles,i):
    key = onlyfiles[i]
    new_key = ""
    for j in range(len(key)):
        if(key[j]!='.'):
            new_key += key[j]
        else: break
    new_key+=".xlsx"
    return new_key

# Rozdziela plik z jiry na Users i Worklog, zwraca plik Worklog i zapisuje w folderze
def worklog(ex_file):
    with xw.App(visible=False) as app:
        wb = app.books.open(ex_file)
        for sheet in wb.sheets:
            wb_new = app.books.add()
            sheet.copy(after=wb_new.sheets[0])
            wb_new.sheets[0].delete()
            if sheet.name == "Worklog":
                wb_new.save(f'{sheet.name}.xlsx')
                wb_new.close()
                break
            else: continue

# zwraca tablice plików w folderze csvresult (sprawdza nazwy projktów)
def check():
    onlyfiles = [f for f in listdir("./") if isfile(join("./",f))]
    n=len(onlyfiles)
    table = []
    for i in range(n):
        if onlyfiles[i] != '.DS_Store':
            table.append(onlyfiles[i])
    return table

def rozdziel(pliki):
    nowe =[]
    pliki.reverse()
    n=len(pliki)
    for i in range(n):
        plik = pliki[i]
        plik = plik.split(".")
        plik=plik[0]
        nowe.append(plik)
    return nowe

def main():
    konsultanci = pd.read_excel('lista_konsultantow.xlsx', index_col=None)
    os.remove('lista_konsultantow.xlsx')
    if os.path.isfile('~$result.xlsx'):
        os.remove('~$result.xlsx')
    if os.path.isfile('~lista_konsultantow.xlsx'):
        os.remove('~$lista_konsultantow.xlsx')

    ex_file=find()
    print(ex_file)
    worklog(ex_file)
    blkprm = pd.read_excel('Worklog.xlsx', index_col=None)
    blkprm.to_csv('blkprm.csv', encoding='utf-8', index=False, header=True)
    os.remove('Worklog.xlsx')

    #R run
    konsultanci.to_csv('kons.csv', encoding='utf-8', index=False, header=True)
    subprocess.call("Rscript /Users/janpoglod/Desktop/coi_projekty/raporty/test_program.R", shell=True)  
    os.remove('blkprm.csv')
    os.remove('kons.csv')
    #znadowanie nazw plików 
    table=is_csv()
    n=len(table)

    # zamiana csv na excel
    for i in range(n):
        napiscsv = f"{table[i]}"
        read_file = pd.read_csv(napiscsv)
        napisxml = to_xlsx(table, i)
        napisxml = f"{napisxml}"
        read_file.to_excel(napisxml, index = False, startcol=0)

    #usuwanie csv
    for i in range(n):
        os.remove(f'{table[i]}')

    os.remove(ex_file)

    # merge wszystkich arkuszy w jeden
    pliki = is_xlsx()
    with xw.App(visible=False) as app:
        combined_wb = app.books.add()
        for excel_file in pliki:
            wb = xw.Book(excel_file)
            for sheet in wb.sheets:
                sheet.copy(after=combined_wb.sheets[0])
            wb.close()

        combined_wb.sheets[0].delete()
        combined_wb.save(f'result.xlsx')
        combined_wb.close()

    # usuwanie niepotrzebnych plików excel
    for i in range(n):
        os.remove(f'{pliki[i]}')

    #stylistyka
    #nazywanie sheets, kolorowanie usuwanie 1. kolumny
    nazwy=rozdziel(pliki)
    print(nazwy)
    ss=openpyxl.load_workbook("result.xlsx")
    sheets = ss.sheetnames
    s=len(sheets)

    for i in range(s):
        ss_sheet = ss[f'{sheets[i]}']
        ss_sheet.title = f'{nazwy[i]}'
        ss.save("result.xlsx")
    #usuniecie 1. wiersza

    for i in range(s):
        ss_sheet = ss[f'{nazwy[i]}']
        ss_sheet.delete_cols(1)
        nrow = ss_sheet.max_row
        fill_cell = PatternFill(patternType='solid', fgColor='1E90FF') 
        for cell in ss_sheet[f"{1}:{1}"]:
                cell.fill = fill_cell
        for j in range(2,nrow+1):
            if j%2==0:
                fill_cell = PatternFill(patternType='solid', fgColor='c9e1f8') 
            else:
                fill_cell = PatternFill(patternType='solid', fgColor='FFFFFF') 
            for cell in ss_sheet[f"{j}:{j}"]:
                cell.fill = fill_cell

    ss.save("result.xlsx")
    konsultanci.to_excel('lista_konsultantow.xlsx', index=False)
    blkprm.to_excel('stara_wersja.xlsx')

main()
